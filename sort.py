from openpyxl import load_workbook
from collections import Counter


def hierarchical_sort(rows, columns):
    counters = [Counter() for _ in columns]
    first_seen = [{} for _ in columns]

    for position, row in enumerate(rows):
        values = tuple(row[column] for column in columns)

        for level in range(len(columns)):
            prefix = values[:level + 1]

            counters[level][prefix] += 1
            first_seen[level].setdefault(prefix, position)

    def get_sort_key(row):
        values = tuple(row[column] for column in columns)
        key = []

        for level in range(len(columns)):
            prefix = values[:level + 1]

            key.append(-counters[level][prefix])
            key.append(first_seen[level][prefix])

        return tuple(key)

    rows.sort(key=get_sort_key)

    return rows, counters


wb = load_workbook("python.xlsx")
ws = wb.active

# min_row=2：略過第一列標題
rows = [
    row
    for row in ws.iter_rows(min_row=2, values_only=True)
    if any(value is not None for value in row[:5])
]

# 按照 A、B、C、D、E 欄階層排序
rows, counters = hierarchical_sort(
    rows,
    columns=(0, 1, 2, 3, 4)
)

# 從第 2 列開始寫回，保留第一列標題
for row_number, row_data in enumerate(rows, start=2):
    for column_number, value in enumerate(row_data, start=1):
        ws.cell(
            row=row_number,
            column=column_number,
            value=value
        )

wb.save("分類結果.xlsx")
