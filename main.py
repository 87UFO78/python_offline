# excel 資料排序
from openpyxl import load_workbook
from collections import Counter

wb = load_workbook('python.xlsx')
ws = wb.active

rows = [
    row
    for row in ws.iter_rows(values_only = True)
    if row[1] is not None
]
'''
# 串列式生成
rows = []

for row in ws.iter_rows(values_only=True):
    if row[1] is not None:
        rows.append(list(row))

'''


result = Counter(row[1] for row in rows)

order = {
    fruit: index
    for index, (fruit, _) in enumerate(result.most_common())
}

"""
order = {}

for index, item in enumerate(result.most_common()):
    fruit = item[0]
    count = item[1]

    order[fruit] = index
"""

rows.sort(key = lambda row: order[row[1]])

"""
# lamba row: order[row[1]] 
def get_order(row):
    fruit = row[1]
    return order[fruit]
"""

for row_number, row_data in enumerate(rows, start = 1):
    for column_number, value in enumerate(row_data, start = 1):
        ws.cell(row = row_number, column = column_number, value = value)

wb.save('python.xlsx')

print(result)
